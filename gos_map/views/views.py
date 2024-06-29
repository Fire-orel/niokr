# views.py

from django.contrib.auth import login,logout
from django.shortcuts import render, redirect,get_object_or_404
from django.views import View
from ldap3 import Server, Connection, ALL, NTLM
from ..forms import LoginForm,MapForms,PublicationForms,SecurityDocumentsForms,MonographsForms,EventForms,GrantForms,NIRSForms,PopularSciencePublicationsForms,ScientificDirectionsForms,InternationalCooperationForms,TypePublicationsForms,TypeDocumentsForms,TypePropertyForms,TypeMonographsForms
from ..models import UserManager,Map,Publications,TypePublications,SecurityDocuments,Monographs,Event,Grant,NIRS,PopularSciencePublications,ScientificDirections,InternationalCooperation,Faculty,Department,TypeDocuments,TypeProperty,TypeMonographs
from django.views.generic import ListView,DetailView,UpdateView
from django.http import JsonResponse,HttpResponseRedirect,HttpResponse
from openpyxl import Workbook
import zipfile
import io
import os
from dotenv import load_dotenv
load_dotenv()



class LoginView(View):
    form_class = LoginForm
    template_name = 'login.html'

    def get(self, request):
        form = self.form_class()
        return render(request, self.template_name, {'form': form})

    def post(self, request):
        form = self.form_class(request.POST)
        if form.is_valid():
            username = form.cleaned_data['username']
            password = form.cleaned_data['password']

            # Сначала проверяем пользователя в базе данных Django
            try:
                user = UserManager.get_user_by_login(login=username)
                if user !=False :
                    if UserManager.get_user_by_login_password(login=username,password=password):
                        request.session['user_id'] = user.id
                        request.session['user_full_name'] = user.full_name
                        return redirect('home')
                    else:
                        return render(request, self.template_name, {'form': form, 'error': 'Логин или пароль неверны'})
            except UserManager.DoesNotExist:
                user = None

            # Если пользователь не найден в базе данных Django, пробуем аутентификацию через LDAP
            ldap_user = self.ldap_authenticate(username, password)

            if ldap_user:
                # Создаем пользователя в базе данных Django
                first_name=ldap_user['first_name']
                last_name=ldap_user['last_name']
                if first_name is None:
                    first_name=""
                if last_name is None:
                    last_name=""
                user = UserManager.objects.create(
                    login=ldap_user['username'],
                    password=password,  # Здесь пароль лучше захэшировать

                    full_name=first_name+' '+ last_name

                )
                user.save()

                login(request, user, backend='django.contrib.auth.backends.ModelBackend')
                request.session['user_id'] = user.pk
                request.session['user_full_name'] = user.full_name

                return redirect('user_detail')

            # Если аутентификация не удалась, выводим сообщение об ошибке
            return render(request, self.template_name, {'form': form, 'error': 'Логин или пароль неверны'})

        return render(request, self.template_name, {'form': form})

    def ldap_authenticate(self, username, password):
        LDAP_SERVER = os.getenv("server_adress")
        LDAP_USER = f'{os.getenv('server_domen')}\\{username}'
        LDAP_PASSWORD = password
        LDAP_SEARCH_BASE = 'DC=FIREOREL,DC=ru'

        # NTLM поддержка через ldap3
        server = Server(LDAP_SERVER, get_info=ALL)
        conn = Connection(server, user=LDAP_USER, password=LDAP_PASSWORD, authentication=NTLM)

        if not conn.bind():
            return None

        conn.search(LDAP_SEARCH_BASE, f'(sAMAccountName={username})', attributes=['sAMAccountName', 'givenName', 'sn'])
        if len(conn.entries) != 1:
            return None


        user_info = conn.entries[0]

        return {
            'username': user_info.sAMAccountName.value,
            'first_name': user_info.givenName.value,
            'last_name': user_info.sn.value,
        }

class UserDetail(View):
    template_name = 'user_detail.html'
    def get(self, request):
        user_full_name = request.session.get('user_full_name')
        position = UserManager.get_position_id(request.session.get("user_id"))

        context = {
            'user_full_name': user_full_name,
            'user_id':UserManager.get_user_id(request.session.get("user_id")),
            'user_posistion':UserManager.get_user_id(request.session.get("user_id")).position,
            'facultys':Faculty.objects.all(),
            'departments':Department.objects.all(),


        }
        return render(request, self.template_name, context)

    def post(self, request):

        if request.POST.get('logout'):
            request.session.flush()
            return redirect('login')

        user_id=request.session.get("user_id")
        position = request.POST.get('position')
        faculty = request.POST.get('faculty')
        department=request.POST.get('department')


        user=UserManager.objects.get(pk=user_id)
        if position=='НО':
            user.position=position
        elif position=='ЗД':
            user.position=position
            user.faculty=faculty
        elif position=='ЗК':
            user.position=position
            user.department=department

        user.save()

        return redirect('home')






class HomeView(View):
    template_name = 'home.html'

    def get(self, request):
        user_full_name = request.session.get('user_full_name')
        position = UserManager.get_position_id(request.session.get("user_id"))
        user_id=request.session.get("user_id")
        faculty=UserManager.objects.get(pk=user_id).faculty
        maps=Map.objects.none()
        faculty_list=Faculty.objects.none()
        department_list=Department.objects.none()

        if  position =="НО":
            maps= Map.objects.all ()
            departments_list=Department.objects.all()
            faculty_list=Faculty.objects.all()
        elif  position =="ЗД":
            departments=Department.objects.filter(faculty=faculty)
            departments_list=Department.objects.filter(faculty=faculty)
            for department in departments:
                maps=maps|Map.objects.filter(department=department)
        elif position =="ЗК":
            departmentss=UserManager.objects.get(pk=user_id).department
            departments_list=Department.objects.filter(name_department=departmentss)

            maps = Map.objects.filter(responsible=UserManager.get_user_id(request.session.get("user_id")))



        context = {
            'user_full_name': user_full_name,
            'user_id':UserManager.get_user_id(request.session.get("user_id")),
            'user_posistion':UserManager.get_user_id(request.session.get("user_id")).position,
            'maps': maps,
            'mapform':MapForms(user=UserManager.objects.get(pk=user_id)),
            'type_publications':TypePublications.objects.all(),
            'facultys':faculty_list,
            'departments':departments_list,
            'type_publications':TypePublications.objects.all(),
            'typepublicationsforms':TypePublicationsForms(),
            'type_documents':TypeDocuments.objects.all(),
            'typedocumentsforms':TypeDocumentsForms(),
            'type_propertys':TypeProperty.objects.all(),
            'typepropertyforms':TypePropertyForms(),
            'typemonographs':TypeMonographs.objects.all(),
            'typemonographsforms':TypeMonographsForms()




        }
        return render(request, self.template_name, context)

    def post(self, request):
        # print(request.POST.get('logout'))
        if request.POST.get('logout'):
            request.session.flush()
            return redirect('login')


class CheckMap(View):
    def post(self, request, *args, **kwargs):
        year = request.POST.get("year")
        quarter = request.POST.get("quarter")
        department = request.POST.getlist("department")

        if len(department)>1:
            return JsonResponse({'message': 'Invalid request method'}, status=400)

        department=department[0]

        check=Map.check_data(year,quarter,department)


        # Обработка данных
        if not check:  # Замените на ваше условие
            map=Map.objects.create(
                year=year,
                quarter=quarter,
                department=department,
                comment=request.POST.get("comment"),
                responsible=UserManager.get_user_id(request.session.get("user_id"))
            )
            map.save()
            return JsonResponse({'message': 'Success'}, status=200)
        else:
            return JsonResponse({'message': 'Error'}, status=400)

    def get(self, request, *args, **kwargs):
        return JsonResponse({'message': 'Invalid request method'}, status=400)

class deleteMap(View):
    def post(self, request, pk):
        try:
            map = Map.objects.get(pk=pk)
            map.delete()
            return JsonResponse({'message': 'Map deleted successfully'}, status=200)
        except Map.DoesNotExist:
            return JsonResponse({'error': 'Map not found'}, status=404)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)




class MapDetails(View):
    def get(self,request,pk):
        status="False"

        user_full_name = request.session.get('user_full_name')



        data = get_object_or_404(Map, pk=pk)
        request.session['map_id']=pk
        # print(request.session['map_id'])

        publications=Publications.objects.filter(id_map=Map.get_map_id(pk))
        publications_count=publications.filter(status="Завершено").count()

        securitydocuments = SecurityDocuments.objects.filter(id_map=Map.get_map_id(pk))
        securitydocuments_count=securitydocuments.filter(status="Завершено").count()

        monographs=Monographs.objects.filter(id_map=Map.get_map_id(pk))
        monographs_count=monographs.filter(status="Завершено").count()

        events=Event.objects.filter(id_map=Map.get_map_id(pk))
        events_count=events.filter(status="Завершено").count()

        grant=Grant.objects.filter(id_map=Map.get_map_id(pk))
        grant_count=grant.filter(status="Завершено").count()

        nirs = NIRS.objects.filter(id_map=Map.get_map_id(pk))
        nirs_count=nirs.filter(status="Завершено").count()

        popularsciencepublications = PopularSciencePublications.objects.filter(id_map=Map.get_map_id(pk))
        popularsciencepublications_count=popularsciencepublications.filter(status="Завершено").count()

        scientificdirections=ScientificDirections.objects.filter(id_map=Map.get_map_id(pk))
        scientificdirections_count=scientificdirections.filter(status="Завершено").count()

        internationalcooperation=InternationalCooperation.objects.filter(id_map=Map.get_map_id(pk))
        internationalcooperation_count=internationalcooperation.filter(status="Завершено").count()



        if len(publications)==publications_count and len(securitydocuments)==securitydocuments_count and len(monographs)==monographs_count and len(events)==events_count and len(grant)==grant_count and len(nirs)==nirs_count and len(popularsciencepublications)==popularsciencepublications_count and len(scientificdirections)==scientificdirections_count and len(internationalcooperation)==internationalcooperation_count:

            status="True"



        context = {
            'status':status,

            'user_full_name': user_full_name,
            'map':data,

            'publications':publications,
            'publicationsforms':PublicationForms(),


            'securitydocuments':securitydocuments,
            'securitydocumentsforms':SecurityDocumentsForms(),

            'monographs':monographs,
            'monographsforms':MonographsForms(),

            'events':events,
            'eventforms':EventForms(),

            'grants':grant,
            'grantforms':GrantForms(),

            'nirss':nirs,
            'nirsforms':NIRSForms(),

            'popularsciencepublicationss':popularsciencepublications,
            'popularsciencepublicationsforms':PopularSciencePublicationsForms(),
            'scientificdirectionss':scientificdirections,
            'scientificdirectionsforms':ScientificDirectionsForms(),
            'internationalcooperations':internationalcooperation,
            "internationalcooperationforms":InternationalCooperationForms(),




        }
        return render(request, 'map_details.html',context)

    def post(self, request,**kwargs):

        if request.POST.get('logout'):
            request.session.flush()
            return redirect('login')

class mapСompleted(View):
    def get(self, request,pk, *args, **kwargs):
        map=get_object_or_404(Map,id=pk)
        map.status="Завершено"
        map.save()
        return redirect('home')

class mapReturn(View):
    def get(self, request,pk, *args, **kwargs):
        map=get_object_or_404(Map,id=pk)
        map.status="Редактируется"
        map.save()
        return redirect('home')




class otchet(View):
    def post(self,request):
        # faculty = request.POST.getlist('faculty')
        department = request.POST.getlist('department')
        quarter = request.POST.getlist('quarter')
        value=request.POST.getlist('hidden_selected_value')
        tables=request.POST.getlist('table_name')
        year = request.POST.getlist('year')

        if len(tables)==0:
            tables=value

        # print(quarter,table,type_table_publication)
        # print(UserManager.get_user_id(request.session.get("user_id")))
        user_id=UserManager.get_user_id(request.session.get("user_id"))
        zip_buffer = io.BytesIO()

        map_2=Map.objects.none()
        map_1=Map.objects.none()
        map=Map.objects.none()

        if len(quarter)>0:
            for i in quarter:
                map=map|Map.objects.filter(quarter=i)
        else :
            map=Map.objects.all()


        if len(department)>0:
            for q in department:
                map_1=map_1|map.filter(department=Department.objects.get(pk=q).name_department)
        else:
            map_1=map


        if len(year)>0:
            for q in year:
                map_2=map_2|map_1.filter(year=q)

        else:
            map_2=map_1

        with zipfile.ZipFile(zip_buffer, 'w') as zip_file:

            for id_map in map_2:

                workbook = Workbook()
                workbook.remove(workbook.active)
                for table in tables:
                    if table=='Publications':
                        count=2
                        sheet = workbook.create_sheet('Публикации')
                        sheet['A1']="Тип публикации"
                        sheet['B1']="ФИО автора"
                        sheet['C1']="Наименование публикации"
                        sheet['D1']="Выходные данные публикации (Название журнала, Номер, Том, страницы)"
                        sheet['E1']="Год"
                        sheet['F1']="Место опубликования"
                        sheet['G1']="Объем публикации (п.л.)"
                        sheet['H1']="eLIBRARY ID"
                        sheet['I1']="DOI публикации"

                        publicationss=Publications.objects.filter(id_map=id_map)
                        for publications in publicationss:
                            sheet[f"A{count}"]=publications.type_publication.name_type_publications
                            sheet[f"B{count}"]=publications.full_name_author_publications
                            sheet[f"C{count}"]=publications.name_publication_publications
                            sheet[f"D{count}"]=publications.exit_data
                            sheet[f"E{count}"]=publications.year
                            sheet[f"F{count}"]=publications.place_publication_publications
                            sheet[f"G{count}"]=publications.volume_publication
                            sheet[f"H{count}"]=publications.eLIBRARY_ID
                            sheet[f"I{count}"]=publications.doi_publication
                            count+=1
                    if table=='SecurityDocuments':
                        count=2
                        sheet = workbook.create_sheet('Охранные документы')
                        sheet['A1']="Тип документа"
                        sheet['B1']="Вид интеллектуальной собственности"
                        sheet['C1']="ФИО автора (ов)"
                        sheet['D1']="Наименование"
                        sheet['E1']="Номер заявки / патента"

                        securitydocumentss=SecurityDocuments.objects.filter(id_map=id_map)
                        for securitydocuments in securitydocumentss:

                            sheet[f"A{count}"]=str(securitydocuments.type_document)
                            sheet[f"B{count}"]=str(securitydocuments.type_property)
                            sheet[f"C{count}"]=securitydocuments.full_name_author_security_documents
                            sheet[f"D{count}"]=securitydocuments.name_publication_security_documents
                            sheet[f"E{count}"]=securitydocuments.application_number

                            count+=1
                    if table=='Monographs':
                        count=2
                        sheet = workbook.create_sheet('Охранные документы')
                        sheet['A1']="Тип монографии (выпадающий список)"
                        sheet['B1']="Автор(ы) ФИО (полностью)"
                        sheet['C1']="Название работы"
                        sheet['D1']="Тираж"
                        sheet['E1']="Объем"
                        sheet['F1']="Издательство(наименование)"
                        sheet['G1']="Вид издательства"
                        sheet['H1']="Год издания"

                        monographss=Monographs.objects.filter(id_map=id_map)
                        for monographs in monographss:

                            sheet[f"A{count}"]=str(monographs.type_monographs)
                            sheet[f"B{count}"]=monographs.full_name_author_monographs
                            sheet[f"C{count}"]=monographs.name_works
                            sheet[f"D{count}"]=monographs.circulation
                            sheet[f"E{count}"]=monographs.volume_monographs
                            sheet[f"F{count}"]=monographs.publishing_house
                            sheet[f"G{count}"]=monographs.type_publishing_house
                            sheet[f"H{count}"]=monographs.year_of_publication_monographs

                            count+=1
                    if table=='Grant':
                        count=2
                        sheet = workbook.create_sheet('Грант')
                        sheet['A1']="Тип гранта (выпадающий список)"
                        sheet['B1']="Наименование фонда"
                        sheet['C1']="Наименование конкурса"
                        sheet['D1']="Код конкурса"
                        sheet['E1']="Номинация (при наличии)"
                        sheet['F1']="Наименование темы проекта (без кавычек)"
                        sheet['G1']="Руководитель проекта (ФИО полностью)"
                        sheet['H1']="Численность проектного коллектива"
                        sheet['I1']="Численность молодых ученых"
                        sheet['J1']="ФИО (полностью) исполнителей"
                        sheet['K1']="Выиграно (Да, Нет)"

                        grants=Grant.objects.filter(id_map=id_map)
                        for grant in grants:

                            sheet[f"A{count}"]=str(grant.type_grant)
                            sheet[f"B{count}"]=grant.name_fund
                            sheet[f"C{count}"]=grant.name_competition
                            sheet[f"D{count}"]=grant.kod_competition
                            sheet[f"E{count}"]=grant.nomination
                            sheet[f"F{count}"]=grant.name_project_topic
                            sheet[f"G{count}"]=grant.project_manager
                            sheet[f"H{count}"]=grant.number_project_team
                            sheet[f"I{count}"]=grant.number_young_scientists
                            sheet[f"J{count}"]=grant.full_name_performer
                            sheet[f"K{count}"]=grant.winner

                            count+=1

                    if table=='Event':
                        count=2
                        sheet = workbook.create_sheet('Мероприятия')
                        sheet['A1']="Тип участия"
                        sheet['B1']="ФИО (полностью) участников ЗабГУ"
                        sheet['C1']="Название мероприятия"
                        sheet['D1']="Уровень (международное, всероссийское и др.)"
                        sheet['E1']="Тип мероприятия"
                        sheet['F1']="Наименование Доклада на мероприятия (для конференций)"
                        sheet['G1']="Дата проведения"
                        sheet['H1']="Место проведения"
                        sheet['I1']="Общее кол-во участников мероприятия"
                        sheet['J1']="Кол-во зарубежных участников мероприятия"
                        sheet['K1']="Кол-во экспонатов (для выставки)"
                        sheet['L1']="С изданием сборника (для конференций)"
                        sheet['M1']="Полученные награды НПР и студентами ЗабГУ (медали, дипломы, грамоты)"
                        sheet['N1']="Ссылка на программу мероприятия"

                        events=Event.objects.filter(id_map=id_map)
                        for event in events:

                            sheet[f"A{count}"]=str(event.type_participation)
                            sheet[f"B{count}"]=event.full_name_author_event
                            sheet[f"C{count}"]=event.name_event_event
                            sheet[f"D{count}"]=event.level
                            sheet[f"E{count}"]=str(event.type_event)
                            sheet[f"F{count}"]=event.title_report
                            sheet[f"G{count}"]=event.date_event_event
                            sheet[f"H{count}"]=event.place_event
                            sheet[f"I{count}"]=event.number_participants
                            sheet[f"J{count}"]=event.number_foreign_participants
                            sheet[f"K{count}"]=event.number_exhibits
                            sheet[f"L{count}"]=event.publication_collection
                            sheet[f"M{count}"]=event.awards
                            sheet[f"N{count}"]=event.link

                            count+=1

                    if table=='NIRS':
                        count=2
                        sheet = workbook.create_sheet('НИРС')
                        sheet['A1']="Численность студентов"
                        sheet['B1']="ФИО студентов"
                        sheet['C1']="Форма участия"
                        sheet['D1']="Наименование мероприятия"
                        sheet['E1']="ФИО научного руководителя"
                        sheet['F1']="Награды/Дипломы"
                        sheet['G1']="Дата мероприятия"


                        nirss=NIRS.objects.filter(id_map=id_map)
                        for nirs in nirss:

                            sheet[f"A{count}"]=nirs.number_students
                            sheet[f"B{count}"]=nirs.full_name_students
                            sheet[f"C{count}"]=str(nirs.form_participation)
                            sheet[f"D{count}"]=nirs.name_event_nirs
                            sheet[f"E{count}"]=nirs.full_name_scientific_supervisor
                            sheet[f"F{count}"]=nirs.awards_diplomas
                            sheet[f"G{count}"]=nirs.date_event_nirs


                            count+=1

                    if table=='PopularSciencePublications':
                        count=2
                        sheet = workbook.create_sheet('Научно-популярные издания')
                        sheet['A1']="Ф.И.О. автора (полностью)"
                        sheet['B1']="Название публикации"
                        sheet['C1']="Место опубликования, издательство, год, № выпуска"
                        sheet['D1']="Объем публикации"
                        sheet['E1']="Примечание"


                        popularsciencepublicationss=PopularSciencePublications.objects.filter(id_map=id_map)
                        for popularsciencepublications in popularsciencepublicationss:

                            sheet[f"A{count}"]=popularsciencepublications.full_name_author
                            sheet[f"B{count}"]=popularsciencepublications.name_publication_popular_science_publications
                            sheet[f"C{count}"]=popularsciencepublications.place_publication_popular_science_publications
                            sheet[f"D{count}"]=popularsciencepublications.volume_popular_science_publications
                            sheet[f"E{count}"]=popularsciencepublications.note


                            count+=1



                    if table=='ScientificDirections':
                        count=2
                        sheet = workbook.create_sheet('Научные направления')
                        sheet['A1']="Научное направление"
                        sheet['B1']="Название научной школы"
                        sheet['C1']="Ведущие ученые в данной области (1-3 человека)"
                        sheet['D1']="Кол-во защищенных докторских диссертаций"
                        sheet['E1']="Кол-во защищенных кандидатских диссертаций"
                        sheet['F1']="Кол-во монографий"
                        sheet['G1']="Кол-во статей WoS/Scopus"
                        sheet['H1']="Кол-во статей ВАК"
                        sheet['I1']="Кол-во статей РИНЦ"
                        sheet['J1']="Кол-во заявок на изобретения"
                        sheet['K1']="Кол-во полученных охранных документов"
                        sheet['L1']="Кол-во организованных международных и (или) всероссийских научных и (или) научно-практических мероприятий"
                        sheet['M1']="Объем финансирования научных исследований (в тыс. рублей): фундаментальных, прикладных, разработок"


                        scientificdirectionss=ScientificDirections.objects.filter(id_map=id_map)
                        for scientificdirections in scientificdirectionss:

                            sheet[f"A{count}"]=scientificdirections.name_scientific_direction
                            sheet[f"B{count}"]=scientificdirections.name_scientific_school
                            sheet[f"C{count}"]=scientificdirections.leading_scientists
                            sheet[f"D{count}"]=scientificdirections.number_defended_doctoral_dissertations
                            sheet[f"E{count}"]=scientificdirections.number_defended_PhD_theses
                            sheet[f"F{count}"]=scientificdirections.number_monographs
                            sheet[f"G{count}"]=scientificdirections.number_articles_WoS_Scopus
                            sheet[f"H{count}"]=scientificdirections.number_articles_VAK
                            sheet[f"I{count}"]=scientificdirections.number_articles_RIHC
                            sheet[f"J{count}"]=scientificdirections.number_applications_inventions
                            sheet[f"K{count}"]=scientificdirections.number_security_documents_received
                            sheet[f"L{count}"]=scientificdirections.number_organized
                            sheet[f"M{count}"]=scientificdirections.amount_funding


                            count+=1

                    if table=='InternationalCooperation':
                        count=2
                        sheet = workbook.create_sheet('Международное сотрудничество ')
                        sheet['A1']="Наименование научно-исследовательских  и научно-технических проектов "
                        sheet['B1']="Наименование научных центров, лабораторий, других подразделений, созданнх с участвием международных или иностранных организаций"
                        sheet['C1']="Наименование тем, проводимых совместных научных исследований "
                        sheet['D1']="Наименование тем научных исследований, опытно-конструкторских работ, проводимых по заказам международных и иностранных организаций"
                        sheet['E1']="Наименование научных программ, разработка и реализация которых осуществляется  совместно с международными или иностранными организациями"


                        internationalcooperations=InternationalCooperation.objects.filter(id_map=id_map)
                        for internationalcooperation in internationalcooperations:

                            sheet[f"A{count}"]=internationalcooperation.name_scientific_research
                            sheet[f"B{count}"]=internationalcooperation.name_scientific_centers
                            sheet[f"C{count}"]=internationalcooperation.name_topics
                            sheet[f"D{count}"]=internationalcooperation.name_research_topics
                            sheet[f"E{count}"]=internationalcooperation.name_scientific_programs



                            count+=1
                    if table=='All':
                        count=2
                        sheet = workbook.create_sheet('Публикации')
                        sheet['A1']="Тип публикации"
                        sheet['B1']="ФИО автора"
                        sheet['C1']="Наименование публикации"
                        sheet['D1']="Выходные данные публикации (Название журнала, Номер, Том, страницы)"
                        sheet['E1']="Год"
                        sheet['F1']="Место опубликования"
                        sheet['G1']="Объем публикации (п.л.)"
                        sheet['H1']="eLIBRARY ID"
                        sheet['I1']="DOI публикации"

                        publicationss=Publications.objects.filter(id_map=id_map)
                        for publications in publicationss:
                            sheet[f"A{count}"]=publications.type_publication.name_type_publications
                            sheet[f"B{count}"]=publications.full_name_author_publications
                            sheet[f"C{count}"]=publications.name_publication_publications
                            sheet[f"D{count}"]=publications.exit_data
                            sheet[f"E{count}"]=publications.year
                            sheet[f"F{count}"]=publications.place_publication_publications
                            sheet[f"G{count}"]=publications.volume_publication
                            sheet[f"H{count}"]=publications.eLIBRARY_ID
                            sheet[f"I{count}"]=publications.doi_publication
                            count+=1

                        count=2
                        sheet = workbook.create_sheet('Охранные документы')
                        sheet['A1']="Тип документа"
                        sheet['B1']="Вид интеллектуальной собственности"
                        sheet['C1']="ФИО автора (ов)"
                        sheet['D1']="Наименование"
                        sheet['E1']="Номер заявки / патента"

                        securitydocumentss=SecurityDocuments.objects.filter(id_map=id_map)
                        for securitydocuments in securitydocumentss:

                            sheet[f"A{count}"]=str(securitydocuments.type_document)
                            sheet[f"B{count}"]=str(securitydocuments.type_property)
                            sheet[f"C{count}"]=securitydocuments.full_name_author_security_documents
                            sheet[f"D{count}"]=securitydocuments.name_publication_security_documents
                            sheet[f"E{count}"]=securitydocuments.application_number

                            count+=1

                        count=2
                        sheet = workbook.create_sheet('Охранные документы')
                        sheet['A1']="Тип монографии (выпадающий список)"
                        sheet['B1']="Автор(ы) ФИО (полностью)"
                        sheet['C1']="Название работы"
                        sheet['D1']="Тираж"
                        sheet['E1']="Объем"
                        sheet['F1']="Издательство(наименование)"
                        sheet['G1']="Вид издательства"
                        sheet['H1']="Год издания"

                        monographss=Monographs.objects.filter(id_map=id_map)
                        for monographs in monographss:

                            sheet[f"A{count}"]=str(monographs.type_monographs)
                            sheet[f"B{count}"]=monographs.full_name_author_monographs
                            sheet[f"C{count}"]=monographs.name_works
                            sheet[f"D{count}"]=monographs.circulation
                            sheet[f"E{count}"]=monographs.volume_monographs
                            sheet[f"F{count}"]=monographs.publishing_house
                            sheet[f"G{count}"]=monographs.type_publishing_house
                            sheet[f"H{count}"]=monographs.year_of_publication_monographs

                            count+=1

                        count=2
                        sheet = workbook.create_sheet('Грант')
                        sheet['A1']="Тип гранта (выпадающий список)"
                        sheet['B1']="Наименование фонда"
                        sheet['C1']="Наименование конкурса"
                        sheet['D1']="Код конкурса"
                        sheet['E1']="Номинация (при наличии)"
                        sheet['F1']="Наименование темы проекта (без кавычек)"
                        sheet['G1']="Руководитель проекта (ФИО полностью)"
                        sheet['H1']="Численность проектного коллектива"
                        sheet['I1']="Численность молодых ученых"
                        sheet['J1']="ФИО (полностью) исполнителей"
                        sheet['K1']="Выиграно (Да, Нет)"

                        grants=Grant.objects.filter(id_map=id_map)
                        for grant in grants:

                            sheet[f"A{count}"]=str(grant.type_grant)
                            sheet[f"B{count}"]=grant.name_fund
                            sheet[f"C{count}"]=grant.name_competition
                            sheet[f"D{count}"]=grant.kod_competition
                            sheet[f"E{count}"]=grant.nomination
                            sheet[f"F{count}"]=grant.name_project_topic
                            sheet[f"G{count}"]=grant.project_manager
                            sheet[f"H{count}"]=grant.number_project_team
                            sheet[f"I{count}"]=grant.number_young_scientists
                            sheet[f"J{count}"]=grant.full_name_performer
                            sheet[f"K{count}"]=grant.winner

                            count+=1


                        count=2
                        sheet = workbook.create_sheet('Мероприятия')
                        sheet['A1']="Тип участия"
                        sheet['B1']="ФИО (полностью) участников ЗабГУ"
                        sheet['C1']="Название мероприятия"
                        sheet['D1']="Уровень (международное, всероссийское и др.)"
                        sheet['E1']="Тип мероприятия"
                        sheet['F1']="Наименование Доклада на мероприятия (для конференций)"
                        sheet['G1']="Дата проведения"
                        sheet['H1']="Место проведения"
                        sheet['I1']="Общее кол-во участников мероприятия"
                        sheet['J1']="Кол-во зарубежных участников мероприятия"
                        sheet['K1']="Кол-во экспонатов (для выставки)"
                        sheet['L1']="С изданием сборника (для конференций)"
                        sheet['M1']="Полученные награды НПР и студентами ЗабГУ (медали, дипломы, грамоты)"
                        sheet['N1']="Ссылка на программу мероприятия"

                        events=Event.objects.filter(id_map=id_map)
                        for event in events:

                            sheet[f"A{count}"]=str(event.type_participation)
                            sheet[f"B{count}"]=event.full_name_author_event
                            sheet[f"C{count}"]=event.name_event_event
                            sheet[f"D{count}"]=event.level
                            sheet[f"E{count}"]=str(event.type_event)
                            sheet[f"F{count}"]=event.title_report
                            sheet[f"G{count}"]=event.date_event_event
                            sheet[f"H{count}"]=event.place_event
                            sheet[f"I{count}"]=event.number_participants
                            sheet[f"J{count}"]=event.number_foreign_participants
                            sheet[f"K{count}"]=event.number_exhibits
                            sheet[f"L{count}"]=event.publication_collection
                            sheet[f"M{count}"]=event.awards
                            sheet[f"N{count}"]=event.link

                            count+=1


                        count=2
                        sheet = workbook.create_sheet('НИРС')
                        sheet['A1']="Численность студентов"
                        sheet['B1']="ФИО студентов"
                        sheet['C1']="Форма участия"
                        sheet['D1']="Наименование мероприятия"
                        sheet['E1']="ФИО научного руководителя"
                        sheet['F1']="Награды/Дипломы"
                        sheet['G1']="Дата мероприятия"


                        nirss=NIRS.objects.filter(id_map=id_map)
                        for nirs in nirss:

                            sheet[f"A{count}"]=nirs.number_students
                            sheet[f"B{count}"]=nirs.full_name_students
                            sheet[f"C{count}"]=str(nirs.form_participation)
                            sheet[f"D{count}"]=nirs.name_event_nirs
                            sheet[f"E{count}"]=nirs.full_name_scientific_supervisor
                            sheet[f"F{count}"]=nirs.awards_diplomas
                            sheet[f"G{count}"]=nirs.date_event_nirs


                            count+=1

                        count=2
                        sheet = workbook.create_sheet('Научно-популярные издания')
                        sheet['A1']="Ф.И.О. автора (полностью)"
                        sheet['B1']="Название публикации"
                        sheet['C1']="Место опубликования, издательство, год, № выпуска"
                        sheet['D1']="Объем публикации"
                        sheet['E1']="Примечание"


                        popularsciencepublicationss=PopularSciencePublications.objects.filter(id_map=id_map)
                        for popularsciencepublications in popularsciencepublicationss:

                            sheet[f"A{count}"]=popularsciencepublications.full_name_author
                            sheet[f"B{count}"]=popularsciencepublications.name_publication_popular_science_publications
                            sheet[f"C{count}"]=popularsciencepublications.place_publication_popular_science_publications
                            sheet[f"D{count}"]=popularsciencepublications.volume_popular_science_publications
                            sheet[f"E{count}"]=popularsciencepublications.note


                            count+=1


                        count=2
                        sheet = workbook.create_sheet('Научные направления')
                        sheet['A1']="Научное направление"
                        sheet['B1']="Название научной школы"
                        sheet['C1']="Ведущие ученые в данной области (1-3 человека)"
                        sheet['D1']="Кол-во защищенных докторских диссертаций"
                        sheet['E1']="Кол-во защищенных кандидатских диссертаций"
                        sheet['F1']="Кол-во монографий"
                        sheet['G1']="Кол-во статей WoS/Scopus"
                        sheet['H1']="Кол-во статей ВАК"
                        sheet['I1']="Кол-во статей РИНЦ"
                        sheet['J1']="Кол-во заявок на изобретения"
                        sheet['K1']="Кол-во полученных охранных документов"
                        sheet['L1']="Кол-во организованных международных и (или) всероссийских научных и (или) научно-практических мероприятий"
                        sheet['M1']="Объем финансирования научных исследований (в тыс. рублей): фундаментальных, прикладных, разработок"


                        scientificdirectionss=ScientificDirections.objects.filter(id_map=id_map)
                        for scientificdirections in scientificdirectionss:

                            sheet[f"A{count}"]=scientificdirections.name_scientific_direction
                            sheet[f"B{count}"]=scientificdirections.name_scientific_school
                            sheet[f"C{count}"]=scientificdirections.leading_scientists
                            sheet[f"D{count}"]=scientificdirections.number_defended_doctoral_dissertations
                            sheet[f"E{count}"]=scientificdirections.number_defended_PhD_theses
                            sheet[f"F{count}"]=scientificdirections.number_monographs
                            sheet[f"G{count}"]=scientificdirections.number_articles_WoS_Scopus
                            sheet[f"H{count}"]=scientificdirections.number_articles_VAK
                            sheet[f"I{count}"]=scientificdirections.number_articles_RIHC
                            sheet[f"J{count}"]=scientificdirections.number_applications_inventions
                            sheet[f"K{count}"]=scientificdirections.number_security_documents_received
                            sheet[f"L{count}"]=scientificdirections.number_organized
                            sheet[f"M{count}"]=scientificdirections.amount_funding


                            count+=1
                        count=2
                        sheet = workbook.create_sheet('Международное сотрудничество ')
                        sheet['A1']="Наименование научно-исследовательских  и научно-технических проектов "
                        sheet['B1']="Наименование научных центров, лабораторий, других подразделений, созданнх с участвием международных или иностранных организаций"
                        sheet['C1']="Наименование тем, проводимых совместных научных исследований "
                        sheet['D1']="Наименование тем научных исследований, опытно-конструкторских работ, проводимых по заказам международных и иностранных организаций"
                        sheet['E1']="Наименование научных программ, разработка и реализация которых осуществляется  совместно с международными или иностранными организациями"


                        internationalcooperations=InternationalCooperation.objects.filter(id_map=id_map)
                        for internationalcooperation in internationalcooperations:

                            sheet[f"A{count}"]=internationalcooperation.name_scientific_research
                            sheet[f"B{count}"]=internationalcooperation.name_scientific_centers
                            sheet[f"C{count}"]=internationalcooperation.name_topics
                            sheet[f"D{count}"]=internationalcooperation.name_research_topics
                            sheet[f"E{count}"]=internationalcooperation.name_scientific_programs



                            count+=1
                file_buffer = io.BytesIO()
                workbook.save(file_buffer)
                file_buffer.seek(0)

                # Добавляем файл в ZIP-архив
                map_date= str(id_map.year)+" "+str(id_map.quarter)+" "+str(id_map.department)

                zip_file.writestr(f'{map_date.__str__()}.xlsx', file_buffer.read())

        response = HttpResponse(zip_buffer.getvalue(), content_type='application/zip')
        response['Content-Disposition'] = 'attachment; filename=tables.zip'

        return response
